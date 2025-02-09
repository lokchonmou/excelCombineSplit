import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def create_excel_with_format(dataframe, filename, sheet_name, index = False, header = True):
    
    # 加載現有的工作簿
    wb = load_workbook(filename)
    
    # 如果工作表存在，則刪除它
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    # 創建一個新的工作表
    ws = wb.create_sheet(title=sheet_name)

    # 將 DataFrame 寫入工作表
    start_row = 0 

    for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=index, header=header), start_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 1, column=c_idx, value=value)

    # 設定字型和框線
    font = Font(name='Microsoft JhengHei', size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='2E5F7F', end_color='2E5F7F', fill_type='solid')
    odd_fill = PatternFill(start_color='C7E5F3', end_color='C7E5F3', fill_type='solid')  # 淺藍色
    even_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')  # 白色

    for r_idx, row in enumerate(ws.iter_rows(), 1):
        for cell in row:
            cell.font = font
            cell.border = border
            if r_idx == 1 and header:
                cell.fill = header_fill
                cell.font = Font(name='Microsoft JhengHei', size=12, bold=True, color='FFFFFF')
            elif index and cell.column == 1:
                cell.fill = header_fill
                cell.font = Font(name='Microsoft JhengHei', size=12, bold=True, color='FFFFFF')
            else:
                cell.fill = odd_fill if r_idx % 2 == 1 else even_fill

    # 自動調整欄寬
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length+4) * 1.3  # 乘以一個系數來調整欄寬
        ws.column_dimensions[column].width = adjusted_width

    # 保存工作簿
    wb.save(filename)